#!/usr/bin/env python

""" Main of the tool getting stats for a strava club """

#TODO: Allow updating (not from scratch) excel sheet. This is needed because only last 200 entries are available
#TODO: Allow agregation of data between two dates (for week and month stats)
#TODO: produce three sorted lists (distance, duration, elevation)
#TODO: add graphs in Excel output file
#TODO: add option in command line to disable update from server (only stats). This could be useful if stats are produced using several calls to the script (first for month then for week stats)
#TODO: What about other activities (swim, bike, ...)
#TODO: Add club id in command line

###############################################################################
# IMPORTS
###############################################################################
import stravalib
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import YELLOW
from optparse import OptionParser
import logging

###############################################################################
# SCRIPT Meta-Data
###############################################################################
__author__ = "Gilles ANTOONS"
__copyright__ = "Copyright 2017, The Alstom Running Band"
__license__ = "GPL"
__version__ = "1.0.0"
__maintainer__ = "The Alstom Running Band"
__email__ = "gilles.antoons@alstom.com"
__status__ = "Development"

###############################################################
# CONSTANT
###############################################################
ALSTOM_RUNNING_BAND_CLUB_ID=196662


##########################################
# main()
##########################################
def main():
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("AlstomClubStats")
    
    # Parse command line arguments 
    output_filename=""
    strava_access_token=""
    try:
        cmdline = OptionParser()
        cmdline.add_option("-o", "--output_file", dest="output_filename",\
                   help="Path and name to output file", \
                   default="sample.xlsx",\
                   metavar="OUT_FILE")
        cmdline.add_option("-a", "--strava_access_token", dest="strava_access_token",\
                   help="Access token (application token) to be requested on strava website,", \
                   default="",\
                   metavar="ACCESS_TOKEN")
        cmdline.add_option("-c", "--strava_club_id", dest="strava_club_id",\
                   help="Identifier of the Strava club to get stats for", \
                   default=ALSTOM_RUNNING_BAND_CLUB_ID,\
                   metavar="STRAVA_CLUB_ID")
        (options, args) = cmdline.parse_args()
        output_filename = options.output_filename
        strava_access_token = options.strava_access_token
        strava_club_id = options.strava_club_id
    except:
        logger.exception("Error parsing command line: ")
        System.exit(-1)

    try:
        logger.info("Opening session...")
        client = stravalib.Client(access_token=strava_access_token)
        logger.info("Getting reference to club...")
        alstom_club=client.get_club(strava_club_id)

        # Ne pas faire club -> members -> activities of each member 
        #    car pas possible de passer de athlete a activities (pas autorise). 
        #        Seulement possible d'avoir actities ou activitiesstats pour utilisateur connecte 
        #        (= lie a auth app id)
        #  ac_members=client.get_club_members(ALSTOM_RUNNING_BAND_CLUB_ID)
        #  for ath in ac_members:
        #      stats = client.get_athlete_stats(ath.id)
        #      for stat in stats:
        #          ...

        logger.info("Initializing the Excel Workbook...")
        wb = Workbook()
        ws_rawdata = wb.active
        ws_rawdata.title="Activities raw data"
        ws_summary_run = wb.create_sheet(title="Run Summary")
        # Style for header
        font = Font(b=True, color="FF0000")
        fill = PatternFill("solid", fgColor=YELLOW)
        al = Alignment(horizontal="center", vertical="center")
        # Write header of Raw data sheets
        ws_rawdata.append(["Athlete Last Name", "Athlete First Name","Date","Type", "Distance (m)","Duration","Elevation (m)"])
        for r in ws_rawdata['A1:G1']:
           for c in r:
               c.alignment = al
               c.font = font
               c.fill = fill
        ws_summary_run.append(["Athlete Name", "Total distance (m)","Total duration","Total elevation (m)"])
        for r in ws_summary_run['A1:D1']:
           for c in r:
               c.alignment = al
               c.font = font
               c.fill = fill
       
        # Init dictionary to collect summary data
        sumrun=dict()
        
        logger.info("Getting list of activities of club members...")
        acts=client.get_club_activities(ALSTOM_RUNNING_BAND_CLUB_ID)
        for act in acts:
            logger.debug("Reading activity " + str(act.name) + "->" + str(act.type) + " " + str(act.distance) + " in " + str(act.moving_time) + str(act.total_elevation_gain))
            ath = act.athlete
            logger.debug("... of " + str(ath.id) + ":"+ str(ath.firstname) + " " + str(ath.lastname))
            ws_rawdata.append([
                      str(ath.lastname), str(ath.firstname),
                      str(act.start_date),
                      str(act.type),
                      float(act.distance),
                      str(act.moving_time),
                      float(act.total_elevation_gain)])
            #TODO: use stravalib constant for activity type
            if str(act.type) == "Run":
                tmpkey=str(ath.lastname)+" " + str(ath.firstname)
                if tmpkey in sumrun:
                    sumrun[tmpkey]['distance'] = sumrun[tmpkey]['distance'] + act.distance
                    sumrun[tmpkey]['moving_time'] = sumrun[tmpkey]['moving_time'] + act.moving_time
                    sumrun[tmpkey]['total_elevation_gain'] = sumrun[tmpkey]['total_elevation_gain'] + act.total_elevation_gain
                else:
                    sumrun[tmpkey]={'distance':act.distance,
                                    'moving_time':act.moving_time,
                                    'total_elevation_gain':act.total_elevation_gain}

        logger.info("Writing Excel sheet with summaries")
        for k in sumrun:
           ws_summary_run.append([k, 
                                  float(sumrun[k]['distance']),
                                  str(sumrun[k]['moving_time']),
                                  float(sumrun[k]['total_elevation_gain'])
                                  ])
                
        logger.info("Saving Workbook to " + output_filename)   
        wb.save(output_filename)
        logger.info("Workbook saved")
    except:
        logger.exception("Error: ")
    finally:
        # Close and flush loggers
        logging.shutdown()

if __name__ == '__main__':
    main()
